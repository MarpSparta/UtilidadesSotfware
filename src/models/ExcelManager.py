from openpyxl import load_workbook

class ExcelManager:
    def __init__(self, path_exel, path_sell, name_sell, fecha, data_list):
        self.path_exel = path_exel
        self.path_sell = path_sell
        self.name_sell = name_sell
        self.fecha = fecha
        self.data_list = data_list
        self.contador_de_celdas_vacias = 0

    def create_new_excel(self):

        # ruta de nuestro archivo
        filesheet = self.path_exel

        # creamos ell obejeto load_workbook
        wb = load_workbook(filesheet)

        # seleccionamos el archivo
        sheet = wb.active

        # nombre del vendedor
        sheet[f'B1'].value = self.name_sell

        for i in range(3, 103):

            celda = sheet[f'B{i}'].value

            print(celda)
            print(i)


            if celda == None:

                self.contador_de_celdas_vacias += 1

                print(f"{self.contador_de_celdas_vacias} dentro del if celda")

                if  self.contador_de_celdas_vacias == 4:

                    print("RECOBROS")

                    sheet[f'B{i}'].value = "RECOBROS"

                elif  self.contador_de_celdas_vacias == 9:

                    print("SUMAS")
                    print(i,'dentro de sumas')

                    sheet[f'A{i}'].value = "Total Saldos:"

                    sheet[f'B{i}'] = f"= SUM(A3:A{i-1})"

                    sheet[f'E{i}'].value = "Total Abonos:"

                    sheet[f'F{i}'] = f"=SUM(F3:F{i-1})"

                    break

        print("Si llego al try exep")
        try:

            wb.save(fr"{self.path_sell}\{self.name_sell}\{self.name_sell}_{self.fecha}.xlsx")
            wb.close()
        except Exception as e:
            print(e)

    def save_exel(self):
        # ruta de nuestro archivo
        filesheet = self.path_exel
        # creamos ell obejeto load_workbook
        wb = load_workbook(filesheet)
        # seleccionamos el archivo
        sheet = wb.active

        contador = 3

        while (True):

            A = sheet[f'A{contador}'].value
            G = sheet[f'G{contador}'].value

            if (A == None and G != "/"):

                sheet[f'A{contador}'].value = self.data_list[0]  # Saldo

                sheet[f'B{contador}'].value = self.data_list[1]  # Nombre del cliente

                sheet[f'C{contador}'].value = self.data_list[2]  # Fecha

                sheet[f'D{contador}'].value = self.data_list[3]  # Fecha de vencimiento

                sheet[f'E{contador}'].value = self.data_list[4]  # No.Doc

                sheet[f'G{contador}'].value = self.data_list[5]  # Vendedor Asignado

                break

            else:

                contador += 1

        try:

            wb.save(self.path_exel)
            wb.close()

        except Exception:
            pass

    def clean_excel_data(self):
        # ruta de nuestro archivo
        filesheet = self.path_exel
        # creamos ell obejeto load_workbook
        wb = load_workbook(filesheet)
        # seleccionamos el archivo
        sheet = wb.active

        for i in range(250):

            if (i > 2):
                sheet[f'A{i}'].value = ""  # Saldo

                sheet[f'B{i}'].value = ""  # Nombre del cliente

                sheet[f'C{i}'].value = ""  # Fecha

                sheet[f'D{i}'].value = ""  # Fecha de vencimiento

                sheet[f'E{i}'].value = ""  # No.Doc

                sheet[f'F{i}'].value = ""

                sheet[f'G{i}'].value = ""  # Vendedor Asignado

        try:

            wb.save(self.path_exel)
            wb.close()

        except Exception:
            pass

    def add_recobro(self):
        # ruta de nuestro archivo
        filesheet = self.path_exel
        # creamos ell obejeto load_workbook
        wb = load_workbook(filesheet)
        # seleccionamos el archivo
        sheet = wb.active

        contador = 3
        espacios = 1

        while (True):

            A = sheet[f'A{contador}'].value

            if (A == None and espacios <= 4):

                sheet[f'G{contador}'].value = "/"

                espacios += 1

            elif (espacios == 5):

                break

            contador += 1

        try:

            wb.save(self.path_exel)
            wb.close()

        except Exception as e:
            pass
